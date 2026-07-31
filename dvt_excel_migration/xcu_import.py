"""Import a DVT report spreadsheet into TofuPilot.

Written for reports that keep several sample readings inside a single cell and
store their plots as embedded images: neither shape survives the native Excel
import, because the column mapper reads one value per cell and never sees the
drawing layer. This script flattens both.

    pip install tofupilot openpyxl
    export TOFUPILOT_API_KEY=...
    python xcu_import.py "DVT report.xlsx" --procedure-id <uuid> \\
        --serial-number SN-0001 --part-number PCB-REV-A

Waveform CSVs exported from the scope are picked up automatically when a
--waveforms directory is passed: a curve then lands as a real multi-dimensional
measurement (per-axis units, validators, comparable across runs) instead of a
screenshot pinned to the run.
"""

from __future__ import annotations

import argparse
import os
import re
import struct
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl
from tofupilot.v2 import TofuPilot

# Row 50 names the temperature block, row 51 the columns; readings start at 52.
BLOCK_ROW = 50
HEADER_ROW = 51
SHEET = "SMPS reg+ripple"

# Each temperature block repeats (value, units, result) at a fixed offset. The
# sample class sits in row 48 above the same columns.
BLOCK_COLUMNS = (18, 23, 28, 33, 38, 43, 48)  # R, W, AB, AG, AL, AQ, AV
VALUE_OFFSET, UNITS_OFFSET, RESULT_OFFSET = 0, 1, 3

# "22.33 - 10.11" is one reading of a dual-rail test, not two measurements.
RAIL_SPLIT = re.compile(r"\s*-\s*")
NUMERIC = re.compile(r"-?\d+(?:[.,]\d+)?")


@dataclass
class Reading:
    """One test row, expanded to one entry per sample."""

    test_number: str
    description: str
    procedure: str
    conditions: str
    sample: str
    values: list[float]
    units: str
    lower: float | None
    upper: float | None
    outcome: str
    plots: list[str] = field(default_factory=list)


def parse_stacked(cell: object) -> list[list[float]]:
    """Split a cell holding one reading per line into per-sample values.

    "22.33 - 10.11\\n22.34 - 10.10\\n22.35 - 10.11" is three samples of a
    two-rail measurement, so it yields [[22.33, 10.11], [22.34, 10.10], ...].
    """
    if cell is None:
        return []
    if isinstance(cell, (int, float)):
        return [[float(cell)]]

    rows: list[list[float]] = []
    for line in str(cell).replace("\r", "\n").split("\n"):
        if not line.strip():
            continue
        rails = [NUMERIC.search(part) for part in RAIL_SPLIT.split(line)]
        values = [float(m.group().replace(",", ".")) for m in rails if m]
        if values:
            rows.append(values)
    return rows


# Letterheads and icons sit in the drawing layer next to the real captures.
# A scope screenshot is a full window grab, so it is wide, roughly landscape,
# and never a small banner. Filtering on pixels rather than file size keeps a
# lightly-compressed plot and drops a detailed logo.
MIN_PLOT_WIDTH = 600
MIN_PLOT_HEIGHT = 300


def png_size(payload: bytes) -> tuple[int, int] | None:
    """Width and height from a PNG's IHDR, without a decoder dependency."""
    if len(payload) < 24 or payload[:8] != b"\x89PNG\r\n\x1a\n":
        return None
    width, height = struct.unpack(">II", payload[16:24])
    return width, height


def is_plot(payload: bytes) -> bool:
    size = png_size(payload)
    if size is None:
        return True  # not a PNG we can measure; let it through rather than lose data
    width, height = size
    return width >= MIN_PLOT_WIDTH and height >= MIN_PLOT_HEIGHT


def extract_images(workbook_path: Path,
                   out_dir: Path) -> dict[int, list[Path]]:
    """Pull embedded PNGs out of the .xlsx and group them by anchor row.

    openpyxl drops images on load, so the drawing XML is read straight from the
    zip. The anchor row is what ties a plot to the test it proves — the whole
    reason these cannot go in as one undifferentiated pile.

    Logos are skipped, and an image repeated across rows is only kept once.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    by_row: dict[int, list[Path]] = {}
    seen: set[str] = set()

    with zipfile.ZipFile(workbook_path) as z:
        rels = {}
        for name in z.namelist():
            if re.fullmatch(r"xl/drawings/_rels/drawing\d+\.xml\.rels", name):
                body = z.read(name).decode("utf8", "ignore")
                rels[name] = re.findall(
                    r'Id="([^"]+)"[^>]*?media/([^"]+)"', body)

        for name in z.namelist():
            if not re.fullmatch(r"xl/drawings/drawing\d+\.xml", name):
                continue
            rel_key = name.replace("drawings/", "drawings/_rels/") + ".rels"
            media = dict(rels.get(rel_key, []))
            body = z.read(name).decode("utf8", "ignore")

            for anchor in re.finditer(
                r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>.*?embed=\"([^\"]+)\"",
                body,
                re.S,
            ):
                row = int(anchor.group(1)) + 1  # xdr rows are 0-based
                target = media.get(anchor.group(2))
                if not target:
                    continue

                payload = z.read(f"xl/media/{target}")
                if not is_plot(payload):
                    continue  # letterhead or icon, not a capture
                if target in seen:
                    continue  # same image anchored on several rows

                seen.add(target)
                dest = out_dir / target
                if not dest.exists():
                    dest.write_bytes(payload)
                by_row.setdefault(row, []).append(dest)

    return by_row


def load_waveform(path: Path) -> tuple[list[float], list[float]]:
    """Read a two-column scope CSV into (time, amplitude)."""
    xs: list[float] = []
    ys: list[float] = []
    for line in path.read_text(errors="ignore").splitlines():
        parts = line.replace(";", ",").split(",")
        if len(parts) < 2:
            continue
        try:
            xs.append(float(parts[0]))
            ys.append(float(parts[1]))
        except ValueError:
            continue  # header or preamble line
    return xs, ys


def read_report(path: Path, images: dict[int, list[Path]]) -> list[Reading]:
    sheet = openpyxl.load_workbook(path, data_only=True)[SHEET]
    readings: list[Reading] = []

    def text(row: int, col: int) -> str:
        value = sheet.cell(row=row, column=col).value
        return "" if value is None else str(value).strip()

    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        description = text(row, 3)  # C: Testdescription
        if not description:
            continue

        limits = [
            sheet.cell(
                row=row,
                column=col).value for col in (
                10,
                12)]  # J, L
        lower, upper = (
            float(v) if isinstance(v, (int, float)) else None for v in limits
        )
        conditions = " ".join(
            filter(None, (text(row, col) for col in range(5, 10)))  # E..I
        )

        for block in BLOCK_COLUMNS:
            # Row 50 gives the temperature, row 48 the sample class.
            temperature = text(BLOCK_ROW, block) or "Room temp."
            sample_class = text(48, block)
            per_sample = parse_stacked(
                sheet.cell(row=row, column=block + VALUE_OFFSET).value
            )
            if not per_sample:
                continue

            units = text(row, block + UNITS_OFFSET).split("\n")[0]
            outcome = text(row, block + RESULT_OFFSET)

            for index, values in enumerate(per_sample, start=1):
                label = f"{sample_class} #{index}" if sample_class else f"#{index}"
                readings.append(
                    Reading(
                        test_number=text(row, 2) or str(row),  # B: Testnr
                        description=description,
                        procedure=text(row, 4),  # D: Measuring procedure
                        conditions=conditions,
                        sample=f"{label} @ {temperature}",
                        values=values,
                        units=units,
                        lower=lower,
                        upper=upper,
                        outcome=outcome or "UNSET",
                        plots=[str(p) for p in images.get(row, [])],
                    )
                )

    return readings


def to_outcome(raw: str) -> str:
    """Their result column is prose ("Same result, Pass", "ToDo")."""
    lowered = raw.strip().lower()
    if "fail" in lowered:
        return "FAIL"
    if "pass" in lowered:
        return "PASS"
    return "UNSET"


def build_phases(readings: list[Reading],
                 waveforms: Path | None) -> list[dict]:
    phases: dict[str, dict] = {}

    for reading in readings:
        phase = phases.setdefault(
            reading.description,
            {
                "name": reading.description,
                "outcome": "PASS",
                "started_at": datetime.now(timezone.utc),
                "ended_at": datetime.now(timezone.utc),
                "docstring": reading.procedure or None,
                "measurements": [],
            },
        )

        outcome = to_outcome(reading.outcome)
        if outcome == "FAIL":
            phase["outcome"] = "FAIL"

        validators = []
        if reading.lower is not None:
            validators.append(
                {"operator": ">=", "expected_value": reading.lower})
        if reading.upper is not None:
            validators.append(
                {"operator": "<=", "expected_value": reading.upper})

        curve = None
        if waveforms:
            candidate = waveforms / \
                f"{reading.test_number}_{reading.sample}.csv"
            if candidate.exists():
                curve = load_waveform(candidate)

        measurement: dict = {
            "name": f"{reading.description} — {reading.sample}",
            "outcome": outcome,
            "docstring": reading.conditions or None,
        }

        if curve:
            # A real curve carries its own axes, so it stays comparable across
            # samples and temperatures rather than being a picture of a result.
            times, amplitudes = curve
            measurement["x_axis"] = {
                "name": "Time", "units": "s", "data": times}
            measurement["y_axis"] = [
                {
                    "name": reading.description,
                    "units": reading.units or "V",
                    "data": amplitudes,
                    "validators": validators or None,
                }
            ]
        else:
            measurement["measured_value"] = reading.values[0]
            measurement["units"] = reading.units or None
            if validators:
                measurement["validators"] = validators

        phase["measurements"].append(measurement)

    return list(phases.values())


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--procedure-id", required=True)
    parser.add_argument("--serial-number", required=True)
    parser.add_argument("--part-number", required=True)
    parser.add_argument(
        "--waveforms",
        type=Path,
        help="Directory of scope CSVs named <testnr>_<sample>.csv",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    plots_dir = args.workbook.parent / "plots"
    images = extract_images(args.workbook, plots_dir)
    readings = read_report(args.workbook, images)
    phases = build_phases(readings, args.waveforms)

    curves = sum(1 for p in phases for m in p["measurements"] if "y_axis" in m)
    print(
        f"{len(readings)} readings across {len(phases)} phases "
        f"({curves} as waveforms, {sum(len(v) for v in images.values())} plots)")

    if args.dry_run:
        for phase in phases:
            print(
                f"  {phase['name']}: {len(phase['measurements'])} measurements")
        return

    now = datetime.now(timezone.utc)
    client_options = {"api_key": os.environ["TOFUPILOT_API_KEY"]}
    if os.environ.get("TOFUPILOT_URL"):
        client_options["server_url"] = os.environ["TOFUPILOT_URL"]

    with TofuPilot(**client_options) as client:
        run = client.runs.create(
            outcome="FAIL" if any(
                p["outcome"] == "FAIL" for p in phases) else "PASS",
            procedure_id=args.procedure_id,
            started_at=now,
            ended_at=now,
            serial_number=args.serial_number,
            part_number=args.part_number,
            phases=phases,
        )
        print(f"run {run.id}")

        # Screenshots ride along on the run; a curve imported as a measurement
        # above already sits on its own test. Each upload is initialize -> PUT
        # to the pre-signed URL -> finalize.
        uploads = []
        for paths in images.values():
            for path in paths:
                blob = Path(path)
                upload = client.attachments.initialize(name=blob.name)
                response = urlopen(
                    Request(
                        upload.upload_url,
                        data=blob.read_bytes(),
                        method="PUT",
                        headers={"Content-Type": "image/png"},
                    )
                )
                response.read()
                client.attachments.finalize(id=upload.id)
                uploads.append(upload.id)

        if uploads:
            client.runs.update(id=run.id, attachments=uploads)
            print(f"{len(uploads)} plots attached")


if __name__ == "__main__":
    main()
